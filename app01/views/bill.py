from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Sum
from django.utils import timezone
from app01 import models
from app01.utils.pagination import Pagination
from app01.utils.form import BillModelForm
from openpyxl import load_workbook


def bill_list(request):
    """ 账单列表 """
    # 查询数据
    queryset = models.Bill.objects.all()

    # # 转换时区
    for bill in queryset:
        bill.created_at = timezone.localtime(bill.created_at)
        bill.updated_at = timezone.localtime(bill.updated_at)

    page_object = Pagination(request, queryset, page_size=20)

    context = {
        "queryset": page_object.page_queryset,
        "page_string": page_object.html()
    }
    return render(request, 'bill_list.html', context)


def bill_upload(request):
    """ Excel文件上传"""

    if request.method == 'POST' and request.FILES.get("exc"):
        try:
            # 1.获取用户上传的文件对象
            file_object = request.FILES.get("exc")

            # 2.对象传递给openpyxl，由openpyxl读取文件的内容
            wb = load_workbook(file_object)
            sheet = wb.worksheets[0]

            # 3.循环获取每一行数据
            for row in sheet.iter_rows(min_row=2):
                id, project_name, department, amount, currency, supplier, confirmed_by = [cell.value for cell in
                                                                                          row[:7]]

                exists = models.Bill.objects.filter(project_name=project_name).exists()
                if not exists:
                    models.Bill.objects.create(project_name=project_name, department=department, amount=amount,
                                               currency=currency, supplier=supplier, confirmed_by=confirmed_by)

            # return redirect('/bill/list/')
            return JsonResponse({'message': '账单文件上传成功'}, status=200)
        except Exception as e:
            return JsonResponse({'error': f'账单文件上传失败: {str(e)}'}, status=500)

    return JsonResponse({'error': '无效的请求'}, status=400)


def bill_pie(request):
    """ 构造饼图的数据 """
    try:
        # 查询并聚合数据
        supplier_amounts = models.Bill.objects.values('supplier').annotate(total_amount=Sum('amount'))

        # 构建 db_data_list
        db_data_list = [
            {"value": entry['total_amount'], "name": entry['supplier']}
            for entry in supplier_amounts
        ]

        return JsonResponse({'status': 200, 'data': db_data_list}, status=200)

    except Exception as e:
        return JsonResponse({'status': 500, 'error': str(e)}, status=500)


def bill_bar(request):
    """ 构造柱状图的数据 """
    try:
        # 查询并聚合数据
        department_amounts = models.Bill.objects.values('department').annotate(total_amount=Sum('amount'))

        # 构建 db_data_list
        db_data_list = [
            {"department": entry['department'], "total_amount": entry['total_amount']}
            for entry in department_amounts
        ]

        return JsonResponse({'status': 200, 'data': db_data_list}, status=200)

    except Exception as e:
        return JsonResponse({'status': 500, 'error': str(e)}, status=500)


def bill_edit(request, nid):
    """ 账单编辑 """
    row_object = models.Bill.objects.filter(id=nid).first()

    if request.method == "GET":
        form = BillModelForm(instance=row_object)
        return render(request, 'bill_edit.html', {"form": form})

    form = BillModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('/bill/list/')

    return render(request, 'bill_edit.html', {"form": form})


def bill_delete(request):
    """ 删除账单 """
    uid = request.GET.get('uid')
    exists = models.Bill.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({"status": 500, 'error': "删除失败，数据不存在。"}, status=500)

    models.Bill.objects.filter(id=uid).delete()
    return JsonResponse({"status": 200, 'success': '删除数据成功'}, status=200)
